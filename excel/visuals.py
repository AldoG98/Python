import pandas as pd
import openpyxl
from openpyxl.chart import (
    LineChart, BarChart, PieChart, ScatterChart,
    Reference, Series, PieChart3D, BubbleChart
)
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Color
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart.axis import DateAxis
import os
from datetime import datetime
import calendar

def apply_cell_styling(ws):
    """Apply consistent styling to worksheet"""
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    # Style headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Style data cells
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

def add_heatmap_formatting(ws, start_col, end_col, start_row, end_row):
    """Add color scale formatting to a range of cells"""
    color_scale_rule = ColorScaleRule(
        start_type='min', start_color='63BE7B',  # Green
        mid_type='percentile', mid_value=50, mid_color='FFEB84',  # Yellow
        end_type='max', end_color='F8696B'  # Red
    )
    cell_range = f"{ws.cell(row=start_row, column=start_col).coordinate}:{ws.cell(row=end_row, column=end_col).coordinate}"
    ws.conditional_formatting.add(cell_range, color_scale_rule)

def create_enhanced_visualizations(input_file):
    """Create comprehensive expense visualizations with deeper insights"""
    print(f"Reading input file: {input_file}")
    
    output_folder = r"C:\Users\aldog\OneDrive\Desktop\excel\visualizations"
    os.makedirs(output_folder, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = os.path.join(output_folder, f'enhanced_expense_analysis_{timestamp}.xlsx')
    
    try:
        # Read and preprocess data
        df = pd.read_excel(input_file)
        df['date'] = pd.to_datetime(df['date'])
        df['month'] = df['date'].dt.month
        df['year'] = df['date'].dt.year
        df['day_of_week'] = df['date'].dt.day_name()
        df['week_of_month'] = df['date'].apply(lambda x: (x.day-1) // 7 + 1)
        
        writer = pd.ExcelWriter(output_file, engine='openpyxl')
        workbook = writer.book

        # 1. Trend Analysis Sheet
        print("Creating detailed trend analysis...")
        trends_sheet = workbook.create_sheet("Trend Analysis")
        
        # Monthly trends with year-over-year comparison
        monthly_trends = df.groupby(['year', 'month'])['amount'].agg([
            'sum', 'count', 'mean', 'min', 'max'
        ]).round(2).reset_index()
        monthly_trends['month_name'] = monthly_trends['month'].apply(lambda x: calendar.month_name[x])
        
        # Write monthly trends
        monthly_trends.to_excel(writer, sheet_name='Trend Analysis', index=False, startrow=1)
        ws = writer.sheets['Trend Analysis']
        apply_cell_styling(ws)
        add_heatmap_formatting(ws, 4, 8, 2, len(monthly_trends)+1)  # Apply heatmap to numeric columns

        # 2. Pattern Analysis Sheet
        print("Creating pattern analysis...")
        pattern_sheet = workbook.create_sheet("Pattern Analysis")
        
        # Day of week analysis
        dow_stats = df.groupby('day_of_week')['amount'].agg([
            'count', 'sum', 'mean', 'std'
        ]).round(2).reset_index()
        
        # Week of month analysis
        wow_stats = df.groupby('week_of_month')['amount'].agg([
            'count', 'sum', 'mean', 'std'
        ]).round(2).reset_index()
        
        # Write pattern analyses
        dow_stats.to_excel(writer, sheet_name='Pattern Analysis', index=False)
        wow_stats.to_excel(writer, sheet_name='Pattern Analysis', startrow=len(dow_stats)+3, index=False)
        
        ws = writer.sheets['Pattern Analysis']
        apply_cell_styling(ws)
        
        # 3. Category Deep Dive
        print("Creating category deep dive...")
        category_sheet = workbook.create_sheet("Category Deep Dive")
        
        # Category trends over time
        category_trends = df.pivot_table(
            index=['year', 'month'],
            columns='category',
            values='amount',
            aggfunc='sum',
            fill_value=0
        ).round(2).reset_index()
        
        # Category correlations
        category_pivot = df.pivot_table(
            index=df['date'].dt.strftime('%Y-%m'),
            columns='category',
            values='amount',
            aggfunc='sum',
            fill_value=0
        )
        category_corr = category_pivot.corr().round(2)
        
        # Write category analyses
        category_trends.to_excel(writer, sheet_name='Category Deep Dive', index=False)
        category_corr.to_excel(writer, sheet_name='Category Deep Dive', 
                             startrow=len(category_trends)+3)
        
        ws = writer.sheets['Category Deep Dive']
        apply_cell_styling(ws)
        
        # 4. Payment Method Intelligence
        print("Creating payment method intelligence...")
        payment_sheet = workbook.create_sheet("Payment Intelligence")
        
        # Payment method usage patterns
        payment_patterns = df.groupby(['payment_method', 'category']).agg({
            'amount': ['count', 'sum', 'mean'],
            'interest_rate': 'first'
        }).round(2)
        
        # Calculate interest costs and projections
        payment_patterns['monthly_interest'] = (
            payment_patterns[('amount', 'sum')] * 
            payment_patterns[('interest_rate', 'first')] / 100 / 12
        ).round(2)
        
        payment_patterns['annual_interest'] = payment_patterns['monthly_interest'] * 12
        
        # Create summary of totals
        total_summary = pd.DataFrame({
            'Metric': [
                'Total Expenses',
                'Total Credit Card Expenses',
                'Total Cash/Debit Expenses',
                'Total Monthly Interest',
                'Projected Annual Interest',
                'Total Cost (Expenses + Annual Interest)'
            ],
            'Amount': [
                df['amount'].sum(),
                df[df['interest_rate'] > 0]['amount'].sum(),
                df[df['interest_rate'] == 0]['amount'].sum(),
                payment_patterns['monthly_interest'].sum(),
                payment_patterns['annual_interest'].sum(),
                df['amount'].sum() + payment_patterns['annual_interest'].sum()
            ]
        })
        
        # Calculate percentage of expenses on credit vs cash/debit
        total_expenses = df['amount'].sum()
        credit_expenses = df[df['interest_rate'] > 0]['amount'].sum()
        payment_distribution = pd.DataFrame({
            'Payment Type': ['Credit Cards', 'Cash/Debit'],
            'Amount': [
                credit_expenses,
                total_expenses - credit_expenses
            ],
            'Percentage': [
                (credit_expenses / total_expenses * 100).round(2),
                ((total_expenses - credit_expenses) / total_expenses * 100).round(2)
            ]
        })
        
        # Write analyses to Excel
        # Write payment patterns
        payment_patterns.to_excel(writer, sheet_name='Payment Intelligence', startrow=1)
        
        # Write total summary
        total_summary.to_excel(writer, sheet_name='Payment Intelligence', 
                             startrow=len(payment_patterns)+5, index=False)
        
        # Write payment distribution
        payment_distribution.to_excel(writer, sheet_name='Payment Intelligence',
                                   startrow=len(payment_patterns)+len(total_summary)+8, 
                                   index=False)
        
        ws = writer.sheets['Payment Intelligence']
        apply_cell_styling(ws)
        
        # Add titles for each section
        ws.cell(row=1, column=1, value="Detailed Payment Method Analysis").font = Font(bold=True, size=12)
        ws.cell(row=len(payment_patterns)+5, column=1, value="Total Cost Summary").font = Font(bold=True, size=12)
        ws.cell(row=len(payment_patterns)+len(total_summary)+8, column=1, 
                value="Payment Method Distribution").font = Font(bold=True, size=12)
        
        # 5. Statistical Insights
        print("Creating statistical insights...")
        stats_sheet = workbook.create_sheet("Statistical Insights")
        
        # Basic statistics by category
        category_stats = df.groupby('category')['amount'].describe().round(2)
        
        # Spending distribution analysis
        spending_dist = pd.qcut(df['amount'], q=5).value_counts().sort_index()
        
        # Write statistical analyses
        category_stats.to_excel(writer, sheet_name='Statistical Insights')
        spending_dist.to_excel(writer, sheet_name='Statistical Insights', 
                             startrow=len(category_stats)+3)
        
        ws = writer.sheets['Statistical Insights']
        apply_cell_styling(ws)
        
        # 6. Raw Data with Enhanced Formatting
        print("Formatting raw data...")
        df.to_excel(writer, sheet_name='Raw Data', index=False)
        ws = writer.sheets['Raw Data']
        apply_cell_styling(ws)
        
        # Adjust column widths in all sheets
        for sheet in workbook.sheetnames:
            ws = workbook[sheet]
            for column_cells in ws.columns:
                length = 0
                for cell in column_cells:
                    try:
                        if cell.value:
                            length = max(length, len(str(cell.value)))
                    except AttributeError:
                        continue  # Skip merged cells
                
                if length and hasattr(column_cells[0], 'column_letter'):
                    column = column_cells[0].column_letter
                    adjusted_width = (length + 2)
                    ws.column_dimensions[column].width = adjusted_width
        
        print("Saving workbook...")
        writer.close()
        print(f"Enhanced visualizations created successfully in: {output_file}")
        
        return output_file
        
    except Exception as e:
        print(f"Error creating visualizations: {str(e)}")
        raise

if __name__ == "__main__":
    input_file = r"C:\Users\aldog\OneDrive\Desktop\excel\expense_report.xlsx"
    try:
        output_file = create_enhanced_visualizations(input_file)
        print("\nCreated the following analysis sheets:")
        print("1. Trend Analysis - Detailed monthly trends with YoY comparison")
        print("2. Pattern Analysis - Day of week and week of month patterns")
        print("3. Category Deep Dive - Category trends and correlations")
        print("4. Payment Intelligence - Payment method usage patterns and interest analysis")
        print("5. Statistical Insights - Detailed statistical analysis of spending")
        print("6. Raw Data - Original data with enhanced formatting")
    except Exception as e:
        print(f"\nError: {str(e)}")